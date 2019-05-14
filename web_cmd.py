import time
import sqlite3
import os
import sys
import re
import struct
import requests
import asyncio
import aiohttp
import queue
import random
import threading
import urllib.parse
from aiohttp import TCPConnector
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from aiohttp import ClientConnectorCertificateError
import requests.packages.urllib3.util.ssl_
requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'ALL'
# from xici import get_proxy
# import ssl
# import requests.packages.urllib3.util.ssl_
# requests.packages.urllib3.util.ssl_.DEFAULT_CIPHERS = 'ALL'
# from aiohttp import ClientConnectionError  # These exceptions related to low-level connection problems.
# from aiohttp import ClientResponseError  # These exceptions could happen after we get response from server.
# from aiohttp import ClientPayloadError  # This exception can only be raised while reading the response payload 
# from aiohttp import ClientError
# from aiohttp import TooManyRedirects
# from requests.exceptions import ConnectionError
# from requests.exceptions import TooManyRedirects
# from requests.exceptions import ChunkedEncodingError
# from requests.exceptions import ReadTimeout
# from requests.exceptions import ConnectTimeout
# from requests.exceptions import Timeout


types = sys.getfilesystemencoding()
sys.dont_write_bytecode = True

USER_AGENTS = [
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; Zune 4.0; InfoPath.3; MS-RTC LM 8; .NET4.0C; .NET4.0E)',
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.2.3) Gecko/20100401 Firefox/3.0.16 (.NET CLR 3.5.30729)',
    'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; en-US) AppleWebKit/125.4 (KHTML, like Gecko, Safari) OmniWeb/v563.57',
    'Mozilla/5.0 (X11; U; Linux i686; en-US) AppleWebKit/534.7 (KHTML, like Gecko) Chrome/7.0.517.24 Safari/534.7',
    'Mozilla/5.0 (X11; U; Linux x86_64; en-US) AppleWebKit/534.16 SUSE/10.0.626.0 (KHTML, like Gecko) Chrome/10.0.626.0 Safari/534.16',
    'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; en-US) AppleWebKit/534.20 (KHTML, like Gecko) Chrome/11.0.672.2 Safari/534.20',
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US) AppleWebKit/534.7 (KHTML, like Gecko) Chrome/7.0.514.0 Safari/534.7',
    'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_3; de-de) AppleWebKit/531.22.7 (KHTML, like Gecko) Version/4.0.5 Safari/531.22.7',
    'Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/534.4 (KHTML, like Gecko) Chrome/6.0.481.0 Safari/534.4',
    'Mozilla/5.0 (Windows; U; Windows NT 6.0; en-GB; rv:1.9.1b4) Gecko/20090423 Firefox/3.5b4 (.NET CLR 3.5.30729)',
    'Mozilla/5.0 (Windows; U; Windows NT 6.0; nb-NO) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.1.16) Gecko/20101130 AskTbPLTV5/3.8.0.12304 Firefox/3.5.16 (.NET CLR 3.5.30729)',
    'Mozilla/5.0 (X11; U; Linux i686; de; rv:1.9.0.14) Gecko/2009082505 Red Hat/3.0.14-1.el5_4 Firefox/3.0.14',
    'Mozilla/5.0 (X11; U; Linux i686; tr-TR; rv:1.9.0.10) Gecko/2009042523 Ubuntu/9.04 (jaunty) Firefox/3.0.10',
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1b2) Gecko/20060821 Firefox/2.0b2',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; fr; rv:1.9.2.2) Gecko/20100316 Firefox/3.6.2 GTB7.0',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; de-AT; rv:1.9.1b2) Gecko/20081201 Firefox/3.1b2',
    'Opera/9.80 (Macintosh; Intel Mac OS X; U; nl) Presto/2.6.30 Version/10.61',
    'Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)',
]


# 获得一个随机的IP地址 如'137.111.132.9'
def random_IP():
    return '.'.join([str(random.randint(0, 255)) for x in range(0, 4)])


class CIDRHelper:

    # 判断IP地址的格式是否正确
    def ip_format_chk(self, ip_str):
        pattern = r"\b(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\b"
        if re.match(pattern, ip_str):
            return True
        else:
            return False

    def mask_len_chk(self, masklen):
        if masklen > 0 and masklen < 32:
            return True
        else:
            return False

    def parse(self, ip, masklen):
        if self.ip_format_chk(ip) is False:
            return "0.0.0.1", "0.0.0.0"
        ips = ip.split(".")
        binip = 0
        for id in ips:
            binip = binip << 8
            binip += int(id)
        mask = (1 << 32) - 1 - ((1 << (32 - masklen)) - 1)
        a, b, c, d = struct.unpack('BBBB', struct.pack('>I', (binip & mask)))
        start = ".".join([str(a), str(b), str(c), str(d)])
        a, b, c, d = struct.unpack(
            'BBBB',
            struct.pack('>I', (binip & mask) + (2 << (32 - masklen - 1)) - 1))
        end = ".".join([str(a), str(b), str(c), str(d)])
        return start, end


def build(start, end=None):
    hosts = []
    if start is not None:
        if end is not None:
            for num in range(ip2num(start), ip2num(end) + 1):
                hosts.append(num2ip(num))
        else:
            tmp = start.split('/')
            ch = CIDRHelper()
            start, end = ch.parse(tmp[0], int(tmp[1]))
            for num in range(ip2num(start), ip2num(end) + 1):
                hosts.append(num2ip(num))
    return hosts


def ip2num(ip):
    lp = [int(x) for x in ip.split('.')]
    return lp[0] << 24 | lp[1] << 16 | lp[2] << 8 | lp[3]


def num2ip(num):
    ip = ['', '', '', '']
    ip[3] = (num & 0xff)
    ip[2] = (num & 0xff00) >> 8
    ip[1] = (num & 0xff0000) >> 16
    ip[0] = (num & 0xff000000) >> 24
    return '%s.%s.%s.%s' % (ip[0], ip[1], ip[2], ip[3])


def check_ippool(ip):
    pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}$')
    pattern1 = re.compile(
        r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\-\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$'
    )
    if pattern.match(ip):
        return True
    elif pattern1.match(ip):
        return True
    return False


class Main():

    def __init__(self):
        self.scan_status = False
        self.LOOP = False
        self.target_list = queue.Queue()
        self.total_target = 0
        self.scanned_target = 0
        self.success_count = 0
        self.res_list = []
        self.print_lock = threading.Lock()
        self.load_lock = threading.Lock()
        # 获取代理ip的对象
        # self.proxy = get_proxy()

        # ------------------------
        self.cover_length = 15
        self.bar_length = 25
        self.task_num = 5  # =======================================================|   设置任务数    |===========
        self.delay = [0.8, 1.4, 1.8, 2.3, 2.6, 3.3]
        self.thread_num = self.thread_current_num = 100  # ==================+======|   设置线程数    |========================================

        # self.lineEdit.setText("192.168.1.1/24")
        # self.lineEdit.setPlaceholderText("192.168.1.1/24:80,443,100-1000 或 192.168.1.1-192.168.1.10:1-10000")
    def reset(self):
        self.scan_status = False
        self.LOOP = False
        self.total_target = 0
        self.scanned_target = 0
        self.success_count = 0
        self.res_list = []
        self.thread_num = self.thread_current_num = 100

    # 从数据库读取数据
    def readFromDatabase(self):
        k = 1
        cwd = os.getcwd()
        # path = os.path.join(cwd, 'data')
        path = cwd
        for each in os.listdir(path):
            if each.startswith('dataBase') and not each.endswith('.xlsx'):
                print(str(k)+')', each)
            k += 1
        x = input('select database > ')
        file = os.listdir(path)[int(x) - 1]
        filepath = os.path.join(path, file)
        conn = sqlite3.connect(filepath)
        curs = conn.cursor()
        curs.execute('select * from result_data')
        for each in curs.fetchall():
            print(each)
        conn.close()

    def show_info_scanning(self):
        i = 0
        dy = ['-', '/', '-']
        # end = False
        count = 0
        while self.scan_status:
            percent = int((self.scanned_target / self.total_target) * 100)
            hashes = '#' * int(percent / 100 * self.bar_length)
            dy_ = dy[i % 3]
            i += 1
            spaces = ' ' * (self.bar_length - len(hashes) - 1)
            gn = self.success_count - count
            if gn > 0:
                # print(gn)
                for each in range(count, self.success_count):
                    self.show_result(self.res_list[each])
            count += gn    
                
            # if end is False:
            if len(hashes) == self.bar_length:
                # end = True
                sys.stdout.write("\rPercent: [%s] %d%% %d/%d " %(hashes, percent, self.scanned_target, self.total_target))
            else:
                sys.stdout.write("\rPercent: [%s] %d%% %d/%d " %(hashes + dy_ + spaces, percent, self.scanned_target, self.total_target))
            sys.stdout.flush()

            time.sleep(0.3)

        print("\n= ---------------")
        global start_time
        print('cost', str(time.time() - start_time) + 's')
        print('success number', self.success_count)
        print('total url', self.total_target)
        self.output_sqlite()

    def input_start(self):
        global start_time
        self.lineEdit = input('input target (example:"192.168.1.1/24:80,443,7000-8000"):\n> ')
        # 有效输入
        if self.lineEdit is not None and self.lineEdit.strip() != '':
            if self._load_target(self.lineEdit.strip().rstrip()):
                self.LOOP = True
                start_time = time.time()
                print("[+] Total %d targets." % self.target_list.qsize())
                print("[+] Scaning...")
                print("= ---------------")
                if self.target_list.qsize() < self.thread_num:
                    thread_num = self.target_list.qsize()
                else:
                    thread_num = self.thread_num

                self.thread_current_num = thread_num

                self.scan_status = True
                p = threading.Thread(target=self.show_info_scanning)
                p.start()

                threads = []
                for i in range(1, thread_num + 1):  # =-=-=-=-======++++++++++++++++++++++++++++++|   创建线程   |+++++++++++
                    thread_loop = asyncio.new_event_loop()
                    t = threading.Thread(target=self._work, args=(thread_loop, i))
                    threads.append(t)
                    t.start()
                
                for thread in threads:
                    thread.join()
            else:
                self.LOOP = False
                print('[-] no port provided in target!')
        else:
            self.LOOP = False
            print('[-] invalid target!')

        # print('\n[+] wait exist...')
        # time.sleep(2)
        self.scan_status = False
        if self.LOOP:
            p.join()
            x = input('> 是否导出文件(y/n):')
            if x == 'y':
                self.output_csv()

    def show_result(self, result):
        print("\r[{}] {}:{}   - {} {}".format(result['status'], result['host'], result['port'], result['title'], ' '*self.cover_length))        

    def show_menu(self):
        print('\n< Web服务状态探测工具 >')
        print(' 1.开始探测')
        print(' 2.查看数据')
        print(' 3.退出程序')
        return input('\n> 输入选项: ')

    def run_start(self):
        while True:
            pos = self.show_menu()
            if pos == '1':
                self.input_start()
                self.reset()
            elif pos == '2':
                self.readFromDatabase()
            elif pos == '3':
                exit()
            elif pos == '0':
                break
            else:
                print('select error!')
        print('[-] exit')        

    def get_hosts(self, host_all):
        # 192.168.1.1/24
        if '/' in host_all and check_ippool(host_all):
            return build(host_all)
        elif '-' in host_all and check_ippool(host_all):
            host_all = host_all.split('-')
            return build(host_all[0], host_all[1])
        else:
            return [host_all]
    
    def get_ports(self, port_all):
        # 80,90,1-100
        if ',' in port_all:
            port_ = port_all.split(',')
            ports = []
            for port in port_:
                if '-' in port:
                    ports += self.get_ports(port)
                else:
                    ports.append(port)
            return ports
        # 1-100
        elif '-' in port_all and ',' not in port_all:
            start, end = port_all.split('-')
            return [str(each) for each in range(int(start), int(end)+1)]
        # all
        elif port_all == 'all':
            return [str(each) for each in range(1, 10001)]
        # common
        elif port_all == 'common':
            common = '80,443,7000-8000'
            return self.get_ports(common)
        # http
        elif port_all == 'http':
            return ['80', '443']
        else:
            return [port_all]

    def _load_target(self, target):
        # http://localhost
        if "http://" in target or "https://" in target:
            protocol, s1 = urllib.parse.splittype(target)
            host, s2 = urllib.parse.splithost(s1)
            host, port = urllib.parse.splitport(host)
            port = port if port is not None and port != 0 else 443 if protocol == 'https' else 80
            print('[+] load target...')
            self._geturl(host, port, target)
            return True
        elif ":" in target:
            _v = target.split(':')
            host_all, port_all = _v[0], _v[1]
            hosts = self.get_hosts(host_all)
            ports = self.get_ports(port_all)
            print('[+] load target...')
            for host in hosts:
                for port in ports:
                    self._geturl(host, port, host)
            return True
        else:
            return False

    # 把一个地址在放进队列之前判断下当前地址的数量是否超出最大值
    def put_target(self, i, target, url):
        if self.target_list.qsize() > 25500000:
            print('地址太多啦，最大不超过 %d! 之后的地址就不探测啦!' % 65536)
            return False
        self.target_list.put((i, target, url))
        return True

    def file_input(self):
        try:
            txt_file=input('file path: ')
            with open(txt_file, 'r') as f:
                # i = 0
                for target in f.readlines():
                    target = target.replace('\r', '').replace(
                        '\n', '').strip().rstrip()
                    if target == '':
                        continue
                    self._load_target(target)
        except:
            pass

    def output_csv(self):
        filename = input('> set file name: ')
        self.tocsv(self.res_list, filename)

    def _work(self, thread_loop, j):  # ----------++++++++++++++++++++++++++++++++++++++++++++++|   _work    |++++++++++
        k = 0
        asyncio.set_event_loop(thread_loop)
        while True:
            tasks = []
            for i in range(self.task_num):  # tasks中的事件控制在10个
                self.load_lock.acquire()
                if self.target_list.qsize() > 0:
                    # 取出队列当前任务
                    tid, target, url = self.target_list.get(timeout=1.0)
                    self.load_lock.release()
                    tasks.append(self.create_task(tid, target, url))
                    # tasks.append(self.create_task(tid, target, url))
                else:
                    self.load_lock.release()
                    break
            if tasks != []:
                k += 1
                # with self.print_lock:
                #     print('[+] loop %d in thread-%d start +' % (k, j))
                thread_loop.run_until_complete(asyncio.gather(*tasks))
                # with self.print_lock:
                #     print('[+] loop %d in thread-%d complete --' % (k, j))
                time.sleep(random.choice(self.delay))
            else:
                break
        thread_loop.close()
        self.thread_current_num -= 1

    async def create_task(self, tid, target, url):
        protocol, s1 = urllib.parse.splittype(url)
        host, s2 = urllib.parse.splithost(s1)
        host, port = urllib.parse.splitport(host)
        port = port if port is not None and port != 0 else 443 if protocol == 'https' else 80

        status, header, title, length = await self._mycurl(url, random_ua=True, random_ip=True)
        if status != 0:
            self.success_count += 1

        data = {
            'id': tid,
            'target': target,
            'host': host,
            'port': port,
            'status': status,
            'title': title,
            'header': header,
            'service': '' if protocol == None else protocol.replace('://', ''),
        }
        self.scanned_target += 1
        if status != 0:
            self.res_list.append(data)

    async def _mycurl(self, url, params=None, **kwargs):
        headers = kwargs.get('headers')
        if headers is None:
            headers = {}
        headers['Accept-Charset'] = 'GB2312,utf-8;q=0.7,*;q=0.7'
        headers['Accept'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
        headers['Accept-Encoding'] = 'gzip, deflate, sdch, br'
        headers['Referer'] = url

        random_ua = kwargs.get('random_ua')
        if random_ua:
            kwargs.pop('random_ua')
            if 'User-Agent' not in headers.keys():
                headers["User-Agent"] = random.choice(USER_AGENTS)

        random_ip = kwargs.get('random_ip')
        if random_ip:
            kwargs.pop('random_ip')
            if 'X-Forwarded-For' not in headers.keys():
                headers['X-Forwarded-For'] = random_IP()

        kwargs.setdefault('headers', headers)
        kwargs.setdefault('timeout', int(5))
        kwargs.setdefault('ssl', False)
        kwargs.setdefault('max_redirects', 10)
        
        # set proxy--------------------------------------------===========================================----------
        # _proxy = self.proxy.get()
        # if _proxy is not None: 
        #     kwargs.setdefault('proxy', _proxy)
        retries = 1
        while retries:
            try:
                # return request(method, url, params=params, **kwargs)
                async with aiohttp.ClientSession(connector=TCPConnector(ssl=False)) as session:
                    async with session.get(url, **kwargs) as res:
                        if res is None and retries != 0:
                            retries -= 1
                            continue
                        elif res is not None:
                            if len(res.history) > 5:  # url has been redirected too many times
                                res = res.history[0]  # set res to the first response
                            try:
                                content = await res.read()
                            except:
                                content = ''

                            content = self.check_redirect(content, url)
                            length = len(str(res.headers)) + len(content)
                            
                            status = res.status
                            headers = res.headers

                            return status, headers, self._dealres(content), length
                        else:
                            return 0, None, None, 0
            except ClientConnectorCertificateError:
                retries -= 1
                if retries == 0:
                    return 0, None, None, 0
            except Exception:
                retries -= 1
                if retries == 0:
                    return 0, None, None, 0
            # except aiohttp.ClientSSLError
            # except ssl.SSLError:
            #     retries -= 1
            #     if retries == 0:
            #         return 0, None, None, 0

    def check_redirect(self, content, url):

        if content == '':
            return ''
        else:  # meta 跳转
            soup = BeautifulSoup(content, 'html.parser')
            meta = soup.find('meta')
            if meta is not None and meta['http-equiv'] == 'refresh':
                try:
                    url1 = url.split(':')[1]
                    url1 = 'http:' + url1
                    url2 = meta['content'].split('=')[1].strip()
                    url = urllib.parse.urljoin(url1, url2)
                    res = requests.get(url, verify=False)
                    return res.text
                except:
                    return content
            else:
                return content

    def _geturl(self, host, port, target):

        if target.startswith('http://') or target.startswith('https://'):
            url = target
            if not self.put_target(self.total_target, target, url):
                return
        else:
            _pro = 'https://' if int(port) == 443 else 'http://'
            url = _pro + host + ":" + port + '/'
            if not self.put_target(self.total_target, target, url):
                return
        self.total_target += 1

    # 解析request返回的结果
    def _dealres(self, content):
        
        if content == '':
            # print('\n[-] read content from response failed')
            title = u"网页内容读取错误".encode('utf-8')
        soup = BeautifulSoup(content, "html.parser")
        if soup != None:
            codes = ['utf-8', 'gbk']
            title = soup.title
            if title == None or title.string == None or title.string == '':
                title = u"网页没有标题".encode('utf-8')
            else:
                title = title.string
            codes.append(types)
            codes = list(set(codes))
            
            if isinstance(title, bytes):
                for j in range(0, len(codes)):
                    try:
                        title = title.decode(codes[j]).strip().replace(
                            "\r", "").replace("\n", "")
                        break
                    except:
                        if j + 1 == len(codes):
                            title = u'网页标题编码错误'

        return title

    def output_sqlite(self):
        # # 判断数据库文件是否存在
        # file_exist = False
        # if os.path.exists('dataBase'):
        #     print('[+] database exist')
        #     file_exist = True

        # 连接数据库
        database_name = 'dataBase' + time.strftime('%m_%d_%y %H.%M.%S')
        conn = sqlite3.connect(database_name)
        curs = conn.cursor()

        title_table = {
            'id': u'序号',
            'target': u'目标',
            'host': u'主机',
            'port': u'端口',
            'status': u'状态',
            'title': u'标题',
            'header': u'返回信息',
            'service': u'服务',
            'is_self': u'是否为自有地址',
            'is_smp': u'若为自有地址，网站是否已添加至SMP',
            'domain': u'IP对应的域名',
            'is_maintenance': u'是否为维护类网站',
            'is_pro': u'若为专业公司地址，请明确归属网站的处置情况',
            'backup_time': u'网站备案时间',
            'business': u'网站归属业务',
            'department': u'网站归属部门',
            'contacts': u'统一接口人',
            'contacts_way': u'联系方式',
        }
        titleList = [
            'id', 'target', 'host', 'port', 'service', 'status', 'title',
            'header', 'is_self', 'is_smp'
        ]
        titleList += [
            _key for _key in title_table.keys() if _key not in titleList
        ]

        titleList = ['id', 'target', 'host', 'port', 'service', 'status', 'title', 'header', 'is_self', 'is_smp', 'domain', 'is_maintenance', 'is_pro', 'backup_time', 'business', 'department', 'contacts', 'contacts_way']
        #if file_exist is False:  # 创建表
        tblcmd = 'create table result_data (id int(4), target char(30), host char(16), port int(4), service char(10),status int(4), title char(50), header char(500), is_self char(5), is_smp char(5), domain char(20), is_maintenance char(5), is_pro char(5), backup_time char(20), business char(50), department char(50), contacts char(20), contacts_way char(20)) '
        curs.execute(tblcmd)
        print('\n[+] Create table sucess.')

        insert_cmd_temp = 'INSERT INTO result_data ({}) VALUES({})'
        print('[+] writing to database...')
        for result in self.res_list:
            if result['status'] != 0:
                add_cmd_key = ""
                add_cmd_value = ""
                values=[]
                for key in titleList:
                    try:
                        res=result[key]
                        if isinstance(res, str) or isinstance(res, int):
                            values.append(res)
                            add_cmd_value = add_cmd_value + '?,'
                        elif res is not None:
                            values.append(str(res))
                            add_cmd_value = add_cmd_value + '?,'
                        elif res is None:
                            continue
                        add_cmd_key = add_cmd_key + key + ','
                    except:
                        pass
                add_cmd_key = add_cmd_key[:-1]
                add_cmd_value = add_cmd_value[:-1]
                insert_cmd = insert_cmd_temp.format(add_cmd_key, add_cmd_value)
                curs.execute(insert_cmd, values)
        conn.commit()  # 保存数据库状态
        conn.close()
        print('[+] Write to database: %s complete.' % database_name)

    def tocsv(self, datalines, filename, key='Mysheet'):
        # self.output_sqlite(datalines)
        print('[+] Export to %s...' % (filename))
        if os.path.isfile(filename):
            book = load_workbook(filename=filename)
            book.remove(book[key])
        else:
            book = Workbook()
            book.remove(book.active)

        if key not in book.sheetnames:
            ws = book.create_sheet(key)
        else:
            ws = book[key]
            # ws = book.get_sheet_by_name(key)
        title_table = {
            'id': u'序号',
            'target': u'目标',
            'host': u'主机',
            'port': u'端口',
            'status': u'状态',
            'title': u'标题',
            'header': u'返回信息',
            'service': u'服务',
            'is_self': u'是否为自有地址',
            'is_smp': u'若为自有地址，网站是否已添加至SMP',
            'domain': u'IP对应的域名',
            'is_maintenance': u'是否为维护类网站',
            'is_pro': u'若为专业公司地址，请明确归属网站的处置情况',
            'backup_time': u'网站备案时间',
            'business': u'网站归属业务',
            'department': u'网站归属部门',
            'contacts': u'统一接口人',
            'contacts_way': u'联系方式',
        }
        titleList = [
            'id', 'target', 'host', 'port', 'service', 'status', 'title',
            'header', 'is_self', 'is_smp'
        ]
        titleList += [
            _key for _key in title_table.keys() if _key not in titleList
        ]
        i = 0
        for key in titleList:
            if key in title_table.keys():
                _key = title_table[key]
            else:
                _key = key
            i += 1
            ws.cell(row=1, column=i).value = _key
        i = 1
        for line in datalines:
            i = i + 1
            for key in line.keys():
                try:
                    if line[key] == None or line[key] == '':
                        ws.cell(
                            row=i, column=titleList.index(key) + 1).value = ""
                    else:
                        ws.cell(
                            row=i,
                            column=titleList.index(key) + 1).value = str(
                                line[key])
                except UnicodeEncodeError:
                    ws.cell(
                        row=i,
                        column=titleList.index(key) + 1).value = u"导出对象错误"
        # file = os.path.join(os.getcwd(), filename)
        filename = filename + '.xlsx'
        book.save(filename)
        print('[+] Exported to %s successful!' % (filename))


if __name__ == '__main__':

    start_time = 0
    app = Main()
    app.run_start()