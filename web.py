#!/usr/bin/env python
# -*- coding: utf-8 -*-



import os
import sys
import re
import struct
import queue
import random
import threading
import requests
import urllib.parse
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from requests import request
# from PyQt5 import QtCore
# from PyQt5 import QtGui
# from PyQt5 import QtWidgets
from PyQt5.QtCore import QRect
from PyQt5.QtCore import pyqtSignal
from PyQt5.QtCore import QThread
from PyQt5.QtCore import QMetaObject
from PyQt5.QtCore import QCoreApplication
from PyQt5.QtWidgets import QProgressBar
from PyQt5.QtWidgets import QLineEdit
from PyQt5.QtWidgets import QPushButton
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QDialog
from PyQt5.QtWidgets import QTableWidget
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QApplication
from PyQt5.QtWidgets import QTableWidgetItem
from requests.exceptions import ConnectionError
from requests.exceptions import TooManyRedirects
from requests.exceptions import ChunkedEncodingError
from requests.exceptions import ReadTimeout
# from requests.exceptions import ConnectTimeout
# from requests.exceptions import Timeout
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

types=sys.getfilesystemencoding()
sys.dont_write_bytecode = True


USER_AGENTS = [
    'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; WOW64; Trident/5.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; Media Center PC 6.0; Zune 4.0; InfoPath.3; MS-RTC LM 8; .NET4.0C; .NET4.0E)',
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36",
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.2.3) Gecko/20100401 Firefox/3.0.16 (.NET CLR 3.5.30729)',
    'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; en-US) AppleWebKit/125.4 (KHTML, like Gecko, Safari) OmniWeb/v563.57',
    'Mozilla/5.0 (X11; U; Linux i686; en-US) AppleWebKit/534.7 (KHTML, like Gecko) Chrome/7.0.517.24 Safari/534.7',
    'Mozilla/5.0 (X11; U; Linux x86_64; en-US) AppleWebKit/534.16 SUSE/10.0.626.0 (KHTML, like Gecko) Chrome/10.0.626.0 Safari/534.16',
    'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_6; en-US) AppleWebKit/534.20 (KHTML, like Gecko) Chrome/11.0.672.2 Safari/534.20',
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US) AppleWebKit/534.7 (KHTML, like Gecko) Chrome/7.0.514.0 Safari/534.7',
    'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_3; de-de) AppleWebKit/531.22.7 (KHTML, like Gecko) Version/4.0.5 Safari/531.22.7',
    'Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/534.4 (KHTML, like Gecko) Chrome/6.0.481.0 Safari/534.4',
    'Mozilla/5.0 (Windows; U; Windows NT 6.0; en-GB; rv:1.9.1b4) Gecko/20090423 Firefox/3.5b4 (.NET CLR 3.5.30729)',
    'Mozilla/5.0 (Windows; U; Windows NT 6.0; nb-NO) AppleWebKit/533.18.1 (KHTML, like Gecko) Version/5.0.2 Safari/533.18.5',
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.1.16) Gecko/20101130 AskTbPLTV5/3.8.0.12304 Firefox/3.5.16 (.NET CLR 3.5.30729)',
    'Mozilla/5.0 (X11; U; Linux i686; de; rv:1.9.0.14) Gecko/2009082505 Red Hat/3.0.14-1.el5_4 Firefox/3.0.14',
    'Mozilla/5.0 (X11; U; Linux i686; tr-TR; rv:1.9.0.10) Gecko/2009042523 Ubuntu/9.04 (jaunty) Firefox/3.0.10',
    'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1b2) Gecko/20060821 Firefox/2.0b2',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; fr; rv:1.9.2.2) Gecko/20100316 Firefox/3.6.2 GTB7.0',
    'Mozilla/5.0 (Windows; U; Windows NT 6.1; de-AT; rv:1.9.1b2) Gecko/20081201 Firefox/3.1b2',
    'Opera/9.80 (Macintosh; Intel Mac OS X; U; nl) Presto/2.6.30 Version/10.61',
    'Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)',
]

def random_IP():
    return '.'.join([str(random.randint(0, 255)) for x in range(0,4)])

class CIDRHelper:
    def ip_format_chk(self, ip_str):
        pattern = r"\b(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\b"
        if re.match(pattern, ip_str):
            return True
        else:
            return False

    def mask_len_chk(self, masklen):
        if masklen > 0 and masklen < 32:
            return True
        else:
            return False

    def parse(self, ip, masklen):
        if False == self.ip_format_chk(ip):
            return "0.0.0.1", "0.0.0.0"
        ips = ip.split(".")
        binip = 0
        for id in ips:
            binip = binip << 8
            binip += int(id)
        mask = (1 << 32) - 1 - ((1 << (32 - masklen)) - 1)
        a, b, c, d = struct.unpack('BBBB', struct.pack('>I', (binip & mask)))
        start = ".".join([str(a), str(b), str(c), str(d)])
        a, b, c, d = struct.unpack('BBBB', struct.pack('>I', (binip & mask) + (2 << (32 - masklen - 1)) - 1))
        end = ".".join([str(a), str(b), str(c), str(d)])
        return start, end

def build(start,end = None):
    hosts = []
    if start != None:
        if end !=None:
            for num in range(ip2num(start),ip2num(end)+1):
                hosts.append( num2ip(num))
        else:
            tmp = start.split('/')
            ch = CIDRHelper()
            start,end = ch.parse(tmp[0],int(tmp[1]))
            for num in range(ip2num(start),ip2num(end)+1):
                hosts.append( num2ip(num))
    return hosts


def ip2num(ip):
    lp = [int(x) for x in ip.split('.')]
    return lp[0] << 24 | lp[1] << 16 | lp[2] << 8 | lp[3]

def num2ip(num):
    ip = ['', '', '', '']
    ip[3] = (num & 0xff)
    ip[2] = (num & 0xff00) >> 8
    ip[1] = (num & 0xff0000) >> 16
    ip[0] = (num & 0xff000000) >> 24
    return '%s.%s.%s.%s' % (ip[0], ip[1], ip[2], ip[3])

def check_ippool(ip):
    pattern = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/\d{1,2}$')
    pattern1 = re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\-\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
    if pattern.match(ip):
        return True
    elif pattern1.match(ip):
        return True
    return False

class Worker(QThread):
    sinOut = pyqtSignal(str)  # 自定义信号，执行run()函数时，从相关线程发射此信号
    def __init__(self, parent=None):
        super(Worker, self).__init__(parent)
        self.working = True
        # self.num = 0

    def __del__(self):
        self.working = False
        self.wait()

    def set_status(self,bool = True):
        self.working = bool

    def run(self):
        while self.working == True:
            # file_str = 'File index {0}'.format(self.num)  # str.format()
            # self.num += 1

            # 发出信号
            file_str = 'test'
            self.sinOut.emit(file_str)

            # 线程休眠2秒
            self.sleep(1)


class MainWindow(QDialog):
    def closeEvent(self,event):
        reply = QMessageBox.question(self,'提示框','是否要退出程序？',QMessageBox.Yes | QMessageBox.No,QMessageBox.No)
        if reply == QMessageBox.Yes:
            app.quit()
        else:
            event.ignore()

class Ui_MainWindow(QWidget):
    def __init__(self):
        self.target_list = queue.Queue()
        self.total_target = 0
        self.scanned_target = 0
        self.res_list = []
        self.is_continue = True
        self.thread_num = self.thread_current_num = 100
        self.print_lock = threading.Lock()
        self.load_lock = threading.Lock()
        self.progress_thread = Worker()
        super(QWidget, self).__init__()

    def setupUi(self, Form):
        Form.setObjectName("Web服务状态探测工具 (探测目标的HTTP/HTTPS协议)_V1.0.0 by allen909")
        Form.resize(962, 682)
        self.tableWidget = QTableWidget(Form)
        self.tableWidget.setGeometry(QRect(10, 50, 941, 591))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(4)
        self.tableWidget.setRowCount(20)
        for i in range(0,20):
            self.tableWidget.setRowHeight(i, 30)

        item = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        self.tableWidget.setColumnWidth(0, 200)
        item = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        self.tableWidget.setColumnWidth(1, 300)
        item = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        self.tableWidget.setColumnWidth(2, 80)
        item = QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        self.tableWidget.setColumnWidth(3, 400)

        self.lineEdit = QLineEdit(Form)
        self.lineEdit.setGeometry(QRect(10, 10, 491, 31))
        self.lineEdit.setObjectName("lineEdit")
        self.pushButton = QPushButton(Form)
        self.pushButton.setGeometry(QRect(510, 10, 101, 28))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QPushButton(Form)
        self.pushButton_2.setGeometry(QRect(620, 10, 101, 28))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QPushButton(Form)
        self.pushButton_3.setGeometry(QRect(730, 10, 101, 28))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QPushButton(Form)
        self.pushButton_4.setGeometry(QRect(840, 10, 101, 28))
        self.pushButton_4.setObjectName("pushButton_4")
        self.progressBar = QProgressBar(Form)
        self.progressBar.setGeometry(QRect(10, 650, 931, 21))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")

        self.retranslateUi(Form)
        QMetaObject.connectSlotsByName(Form)


    def retranslateUi(self, Form):
        _translate = QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Web服务状态探测工具 (探测目标的HTTP/HTTPS协议)_V1.0.0 by allen909"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Form", "目标"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Form", "URL"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("Form", "状态"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("Form", "标题"))
        self.pushButton.setText(_translate("Form", "文件导入"))
        self.pushButton.clicked.connect(self.file_input)
        self.pushButton_2.setText(_translate("Form", "运行"))
        self.pushButton_2.clicked.connect(self.input_start)
        self.pushButton_3.setText(_translate("Form", "结果导出"))
        self.pushButton_3.clicked.connect(self.output_csv)
        self.pushButton_4.setText(_translate("Form", "清空"))
        self.pushButton_4.clicked.connect(self.lineEdit.clear)
        self.pushButton_4.clicked.connect(self.clear_input)

        self.pushButton_3.setEnabled(False)
        self.pushButton_4.setEnabled(False)
        self.progress_thread.sinOut.connect(self.progress_show)

        self.lineEdit.setText("输入框: 192.168.1.1/24 或者 192.168.1.1-192.168.1.10 或者 192.168.1.1:8080 或者 http://baidu.com/index.php，文件导入仅支持 单一目标")
        return

    def progress_show(self):
        progress_num =  int((self.scanned_target / self.total_target) * 100)
        self.progressBar.setProperty("value", progress_num)
        if progress_num == 100:
            self.pushButton.setEnabled(True)
            self.pushButton_2.setText("运行")
            self.pushButton_3.setEnabled(True)
            self.pushButton_4.setEnabled(True)
            self.progress_thread.set_status(False)
            self.lineEdit.clear()

    def clear_input(self):
        self.is_continue = False
        self.total_target = 0
        self.scanned_target = 0
        self.target_list.queue.clear()
        self.res_list.clear()
        self.tableWidget.clearContents()

    def input_start(self):
        if self.lineEdit.text()!=None and self.lineEdit.text().strip()!= '':
            self._load_target(self.lineEdit.text().strip().rstrip())

        if self.pushButton_2.text() == '运行':
            if self.target_list.qsize() == 0:
                reply = QMessageBox.information(self,
                                                          "提示框",
                                                          "没有目标，请输入地址或者导入地址!",
                                                          QMessageBox.Yes | QMessageBox.No)
                return

            print("[+] Total %d targets." % self.target_list.qsize())
            print("[+] Scaning...")
            self.pushButton.setEnabled(False)
            # self.pushButton_2.setEnabled(True)
            self.pushButton_3.setEnabled(False)
            self.pushButton_4.setEnabled(False)
            self.progress_thread.set_status(True)
            if self.target_list.qsize() < self.thread_num:
                thread_num = self.target_list.qsize()
            else:
                thread_num = self.thread_num
            self.is_continue = True
            # self.total_target = self.target_list.qsize()
            self.progress_thread.start()
            self.thread_current_num = thread_num
            for i in range(0, thread_num):
                t = threading.Thread(target=self._work)
                t.setDaemon(True)
                t.start()
            self.pushButton_2.setText("停止")
        else:
            self.is_continue = False
            self.pushButton_2.setText("运行")
            self.progress_thread.set_status(False)
            self.pushButton_3.setEnabled(True)


    def file_input(self):
        try:
            txt_file, ok = QFileDialog.getOpenFileName(None, "选取文件", "./", "Txt files(*.txt)")
            with open(txt_file, 'r') as f:
                # i = 0
                for target in f.readlines():
                    target = target.replace('\r', '').replace('\n', '').strip().rstrip()
                    if target == '':
                        continue
                    if not self.put_target(self.total_target, target): break
                    #self.target_list.put((self.total_target, target))
                    self.tableWidget.setRowCount(self.total_target+1)
                    self.tableWidget.setItem(self.total_target, 0, QTableWidgetItem(target))
                    self.tableWidget.setRowHeight(self.total_target, 30)
                    self.total_target += 1
            if self.total_target < 20:
                self.tableWidget.setRowCount(20)
        except:
            pass

    def output_csv(self):
        filename = QFileDialog.getSaveFileName(self, 'save file', './','xlsx files(*.xlsx)')
        self.tocsv(self.res_list, filename[0] )

    def _work(self):
        while True:
            self.load_lock.acquire()
            if self.target_list.qsize() > 0 and self.is_continue:
                tid, target = self.target_list.get(timeout=1.0)
                self.load_lock.release()
                if target.startswith('http://') or target.startswith('https://'):
                    protocol, s1 = urllib.parse.splittype(target)
                    host, s2 = urllib.parse.splithost(s1)
                    host, port = urllib.parse.splitport(host)
                    port = port if port != None and port != 0 else 443 if protocol == 'https' else 80
                    url = target
                    res = self._mycurl('get', url)
                else:
                    if ":" in target:
                        _v = target.split(':')
                        host, port = _v[0], _v[1]
                    else:
                        host = target
                        port = 0
                    url, host, port, protocol, res = self._geturl(host, port)

                status, header, title, length = self._dealres(res)
                data = {
                    'id': tid,
                    'target': target,
                    'host': host,
                    'port': port,
                    'status': status,
                    'title': title,
                    'header': header,
                    'service': '' if protocol == None else protocol.replace('://', ''),
                }
                self.print_lock.acquire()
                try:
                    print('[%d][%d] %s: %s' % (tid, status, target, title))
                except:
                    print('[%d][%d] %s: %s' % (tid, status, target, u'网页标题编码错误'))
                self.scanned_target += 1
                self.print_lock.release()
                self.tableWidget.setItem(tid, 0, QTableWidgetItem(target))
                self.tableWidget.setItem(tid, 1, QTableWidgetItem(url))
                self.tableWidget.setItem(tid, 2, QTableWidgetItem(str(status)))
                self.tableWidget.setItem(tid, 3, QTableWidgetItem(title))
                self.res_list.append(data)
            else:
                self.load_lock.release()
                break
        self.thread_current_num -= 1

    def _load_target(self,target):

        # http://localhost
        if "http://" in target or "https://" in target:
            if not self.put_target(self.total_target, target):
                return 
            self.tableWidget.setRowCount(self.total_target + 1)
            self.tableWidget.setItem(self.total_target, 0, QTableWidgetItem(target))
            self.total_target += 1
        else:

            # 192.168.111.1/24
            if '/' in target and check_ippool(target):
                # i = 0
                for each in build(target):
                    if not self.put_target(self.total_target, each): break
                    #self.put_target(self.total_target, each)
                    self.tableWidget.setRowCount(self.total_target + 1)
                    self.tableWidget.setItem(self.total_target, 0, QTableWidgetItem(each))
                    self.total_target += 1

            # 192.168.111.1-192.168.111.3
            elif '-' in target and check_ippool(target):
                _v = target.split('-')
                # i = 0
                for each in build(_v[0], _v[1]):
                    if not self.put_target(self.total_target, each): break
                    self.tableWidget.setRowCount(self.total_target + 1)
                    self.tableWidget.setItem(self.total_target, 0, QTableWidgetItem(each))
                    self.total_target += 1

            # 192.168.111.1
            else:
                # target = target[:-1] if target[-1] == '/' else target
                if not self.put_target(self.total_target, target): return
                self.tableWidget.setRowCount(self.total_target + 1)
                self.tableWidget.setItem(self.total_target, 0, QTableWidgetItem(target))
                self.total_target += 1

        # self.tableWidget.setRowCount(self.total_target + 1)

        if self.total_target < 20:
            self.tableWidget.setRowCount(20)
    #
    def put_target(self,i,obj):
        if self.target_list.qsize() > 65536 :
            msg = '地址太多啦，最大不超过 %d! 之后的地址就不探测啦!' % 65536 
            reply = QMessageBox.information(self,
                                                      "提示框",
                                                      msg,
                                                      QMessageBox.Yes | QMessageBox.No)
            return False
        self.target_list.put((i,obj))
        return True

    def _mycurl(self,method, url, params=None, **kwargs):
        headers = kwargs.get('headers')
        if headers == None:
            headers = {}
        headers['Accept-Charset'] = 'GB2312,utf-8;q=0.7,*;q=0.7'
        headers['Accept'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
        headers['Accept-Encoding'] = 'gzip, deflate, sdch, br'
        headers['Referer'] = url

        random_ua = kwargs.get('random_ua')
        if random_ua:
            if 'User-Agent' not in headers.keys():
                headers["User-Agent"] = random.choice(USER_AGENTS)

        random_ip = kwargs.get('random_ip')
        if random_ip:
            if 'X-Forwarded-For' not in headers.keys():
                headers['X-Forwarded-For'] = random_IP()

        kwargs.setdefault('headers', headers)
        kwargs.setdefault('timeout', int(3))
        kwargs.setdefault('verify', False)
        try:
            return request(method, url, params=params, **kwargs)
        except ConnectionError as e:
            return None
        except ReadTimeout as e:
            return None
        except ChunkedEncodingError as e:
            return None
        except TooManyRedirects as e:
            kwargs.setdefault('allow_redirects', False)
            try:
                return request(method, url, params=params, **kwargs)
            except:
                return None
        except Exception as e:
            return None

    def _geturl(self,host, port, params=None, **kwargs):
        for pro in ['http://', "https://"]:
            _port = port if port != None and port != 0 else 443 if pro == 'https' else 80
            _pro = 'https://' if port == 443 else pro
            url = _pro + host + ":" + str(_port) + '/'
            res = self._mycurl('get', url, params, **kwargs)

            if res != None:
                if pro == 'http://' and res.status_code == 400 and 'The plain HTTP request was sent to HTTPS port' in res.text:
                    continue
                return url, host, _port, _pro, res
        return None, host, port, None, None

    def _dealres(self,result):
        if result != None:
            status = result.status_code
            header = result.headers
            try:
                length = int(result.headers['content-length'])
            except:
                length = len(str(result.headers)) + len(result.text)
            soup = BeautifulSoup(result.text, "html.parser")
            if soup != None:
                codes = ['utf-8', 'gbk']
                title = soup.title
                if title == None or title.string == None or title.string == '':
                    title = u"网页没有标题".encode('utf-8')
                else:
                    if result.encoding != None:
                        try:
                            title = title.string.encode(result.encoding)
                            codes.append(result.encoding)
                        except:
                            title = "[Error Code]".encode('utf-8')
                    else:
                        title = title.string
                codes.append(types)
                for j in range(0, len(codes)):
                    try:
                        title = title.decode(codes[j]).strip().replace("\r", "").replace("\n", "")
                        break
                    except:
                        continue
                    finally:
                        if j + 1 == len(codes):
                            title = u'网页标题编码错误'
            else:
                title = u'网页没有标题'
            return status, header, title, length
        else:
            return 0, None, None, 0

    def tocsv(self, datalines, filename, key='Mysheet'):
        print('[+] Export to %s...' % (filename))
        if os.path.isfile(filename):
            book = load_workbook(filename=filename)
            book.remove(book[key])
        else:
            book = Workbook()
            book.remove(book.active)

        if key not in book.sheetnames:
            ws = book.create_sheet(key)
        else:
            ws = book[key]
            # ws = book.get_sheet_by_name(key)
        title_table = {
            'id': u'序号',
            'target': u'目标',
            'host': u'主机',
            'port': u'端口',
            'status': u'状态',
            'title': u'标题',
            'header': u'返回信息',
            'service': u'服务',
            'is_self': u'是否为自有地址',
            'is_smp': u'若为自有地址，网站是否已添加至SMP',
            'domain': u'IP对应的域名',
            'is_maintenance': u'是否为维护类网站',
            'is_pro': u'若为专业公司地址，请明确归属网站的处置情况',
            'backup_time': u'网站备案时间',
            'business': u'网站归属业务',
            'department': u'网站归属部门',
            'contacts': u'统一接口人',
            'contacts_way': u'联系方式',
        }
        titleList = ['id', 'target', 'host', 'port', 'service', 'status', 'title', 'header', 'is_self', 'is_smp']
        titleList += [_key for _key in title_table.keys() if _key not in titleList]
        i = 0
        for key in titleList:
            if key in title_table.keys():
                _key = title_table[key]
            else:
                _key = key
            i += 1
            ws.cell(row=1, column=i).value = _key
        i = 1
        for line in datalines:
            i = i + 1
            for key in line.keys():
                try:
                    if line[key] == None or line[key] == '':
                        ws.cell(row=i, column=titleList.index(key) + 1).value = ""
                    else:
                        ws.cell(row=i, column=titleList.index(key) + 1).value = str(line[key])
                except UnicodeEncodeError as e:
                    ws.cell(row=i, column=titleList.index(key) + 1).value = u"导出对象错误"
        book.save(filename)
        print('{+] Exported to %s successful!' % (filename))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    #dialog = Dialog()
    #MainWindow = QMainWindow()
    MainWindow = MainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    #ui.setupUi(dialog)
    MainWindow.show()
    #dialog.show()
    #sys.exit(app.exec_())
    sys.exit(app.exec_())
